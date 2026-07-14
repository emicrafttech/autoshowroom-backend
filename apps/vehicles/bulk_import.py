"""Parse CSV/Excel dealer inventory files into vehicle create payloads."""

from __future__ import annotations

import csv
import io
import mimetypes
import re
from pathlib import PurePosixPath
from typing import Any
from urllib.parse import urlparse

from django.core.validators import URLValidator
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.exceptions import ValidationError
from rest_framework.serializers import ValidationError as SerializerValidationError

from apps.billing.limits import active_listing_count, get_listing_limit, get_media_limits
from apps.dealers.models import DealerLocation

from .models import Vehicle, VehicleMedia
from .serializers import VehicleSerializer

BULK_UPLOAD_COLUMNS = [
    "make",
    "model",
    "year",
    "trim",
    "priceNgn",
    "mileageKm",
    "transmission",
    "fuel",
    "colour",
    "bodyType",
    "drivetrain",
    "conditionGrade",
    "negotiable",
    "description",
    "vin",
    "videoLink1",
    "videoLink2",
    "videoLink3",
    "imageLink1",
    "imageLink2",
    "imageLink3",
    "imageLink4",
    "imageLink5",
    "imageLink6",
]

COLUMN_ALIASES = {
    "make": {"make", "brand", "manufacturer"},
    "model": {"model"},
    "year": {"year", "model_year", "modelyear"},
    "trim": {"trim", "variant", "edition"},
    "priceNgn": {"pricengn", "price", "price_ngn", "amount", "asking_price"},
    "mileageKm": {"mileagekm", "mileage", "mileage_km", "odometer", "km"},
    "transmission": {"transmission", "gearbox"},
    "fuel": {"fuel", "fuel_type", "fueltype"},
    "colour": {"colour", "color"},
    "bodyType": {"bodytype", "body_type", "body"},
    "drivetrain": {"drivetrain", "drive", "drive_type"},
    "conditionGrade": {"conditiongrade", "condition", "condition_grade"},
    "negotiable": {"negotiable", "is_negotiable"},
    "description": {"description", "notes", "comments"},
    "vin": {"vin", "chassis", "chassis_number", "chassisnumber"},
    **{
        f"videoLink{index}": {
            f"videolink{index}",
            f"video_link_{index}",
            f"video_url_{index}",
        }
        for index in range(1, 4)
    },
    **{
        f"imageLink{index}": {
            f"imagelink{index}",
            f"image_link_{index}",
            f"image_url_{index}",
            f"photo_link_{index}",
            f"photo_url_{index}",
        }
        for index in range(1, 7)
    },
}

FIELD_OPTIONS = {
    "transmission": [choice for choice, _ in Vehicle.Transmission.choices],
    "fuel": [choice for choice, _ in Vehicle.Fuel.choices],
    "bodyType": [choice for choice, _ in Vehicle.BodyType.choices],
    "drivetrain": [choice for choice, _ in Vehicle.Drivetrain.choices],
    "conditionGrade": [choice for choice, _ in Vehicle.ConditionGrade.choices],
    "negotiable": ["yes", "no"],
}

CSV_OPTION_GUIDE = [
    "# Allowed values for fields with fixed options:",
    *[
        f"# {field}: {' | '.join(options)}"
        for field, options in FIELD_OPTIONS.items()
    ],
    "# Delete the sample vehicle rows and add your inventory below the header.",
]

SAMPLE_ROWS = [
    {
        "make": "Toyota",
        "model": "Camry",
        "year": "2020",
        "trim": "XSE",
        "priceNgn": "18500000",
        "mileageKm": "42000",
        "transmission": "automatic",
        "fuel": "petrol",
        "colour": "Black",
        "bodyType": "sedan",
        "drivetrain": "fwd",
        "conditionGrade": "good",
        "negotiable": "yes",
        "description": "Clean tokunbo example row",
        "vin": "",
        "videoLink1": "https://example.com/camry-walkaround-1.mp4",
        "videoLink2": "https://example.com/camry-walkaround-2.mp4",
        "videoLink3": "https://example.com/camry-walkaround-3.mp4",
        "imageLink1": "https://example.com/camry-front.jpg",
        "imageLink2": "https://example.com/camry-rear.jpg",
        "imageLink3": "https://example.com/camry-left.jpg",
        "imageLink4": "https://example.com/camry-right.jpg",
        "imageLink5": "https://example.com/camry-interior.jpg",
        "imageLink6": "https://example.com/camry-dashboard.jpg",
    },
    {
        "make": "Lexus",
        "model": "RX",
        "year": "2019",
        "trim": "350",
        "priceNgn": "32000000",
        "mileageKm": "51000",
        "transmission": "automatic",
        "fuel": "petrol",
        "colour": "White",
        "bodyType": "suv",
        "drivetrain": "awd",
        "conditionGrade": "excellent",
        "negotiable": "no",
        "description": "Sample SUV row — replace with your stock",
        "vin": "",
        "videoLink1": "",
        "videoLink2": "",
        "videoLink3": "",
        "imageLink1": "https://example.com/lexus-rx-front.jpg",
        "imageLink2": "",
        "imageLink3": "",
        "imageLink4": "",
        "imageLink5": "",
        "imageLink6": "",
    },
]


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if normalized in aliases and field not in mapped:
                mapped[field] = index
                break
    missing = [field for field in ("make", "model", "year", "priceNgn", "mileageKm") if field not in mapped]
    if missing:
        raise ValidationError(
            f"Missing required columns: {', '.join(missing)}. "
            f"Expected headers like: {', '.join(BULK_UPLOAD_COLUMNS)}"
        )
    return mapped


def _cell(row: list[Any], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: str, default: bool = True) -> bool:
    if not value:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "negotiable"}


def _parse_int(value: str, field: str) -> int:
    cleaned = re.sub(r"[^\d]", "", value)
    if not cleaned:
        raise ValidationError({field: f"{field} is required."})
    return int(cleaned)


def _media_links_from_row(
    row: list[Any],
    header_map: dict[str, int],
) -> list[dict[str, str]]:
    validate_url = URLValidator(schemes=["http", "https"])
    links: list[dict[str, str]] = []
    for kind, field_prefix, count in (
        (VehicleMedia.Kind.VIDEO, "videoLink", 3),
        (VehicleMedia.Kind.PHOTO, "imageLink", 6),
    ):
        for index in range(1, count + 1):
            field = f"{field_prefix}{index}"
            url = _cell(row, header_map.get(field))
            if not url:
                continue
            try:
                validate_url(url)
            except Exception as exc:
                raise ValidationError({field: "Enter a valid http or https URL."}) from exc
            links.append({"kind": kind, "url": url})
    return links


def _row_to_payload(
    row: list[Any],
    header_map: dict[str, int],
    *,
    defaults: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    get = lambda field: _cell(row, header_map.get(field))
    payload = {
        "make": get("make"),
        "model": get("model"),
        "year": _parse_int(get("year"), "year"),
        "trim": get("trim") or "Base",
        "priceNgn": _parse_int(get("priceNgn"), "priceNgn"),
        "mileageKm": _parse_int(get("mileageKm"), "mileageKm"),
        "transmission": (get("transmission") or "automatic").lower().replace(" ", "_"),
        "fuel": (get("fuel") or "petrol").lower().replace(" ", "_"),
        "colour": get("colour") or "Unknown",
        "bodyType": (get("bodyType") or "sedan").lower().replace(" ", "_"),
        "drivetrain": (get("drivetrain") or "fwd").lower().replace(" ", "_"),
        "conditionGrade": (get("conditionGrade") or "good").lower().replace(" ", "_"),
        "negotiable": _parse_bool(get("negotiable"), True),
        "notes": get("description") or None,
        "vin": get("vin") or None,
        "status": Vehicle.Status.HIDDEN,
        **defaults,
    }
    if not payload["make"] or not payload["model"]:
        raise ValidationError({"make": "make and model are required."})
    return payload, _media_links_from_row(row, header_map)


def _create_linked_media(vehicle: Vehicle, media_links: list[dict[str, str]]) -> None:
    first_photo = None
    for sort_order, item in enumerate(media_links):
        url = item["url"]
        parsed = urlparse(url)
        file_name = PurePosixPath(parsed.path).name or (
            f"{item['kind']}-{sort_order + 1}"
        )
        content_type = mimetypes.guess_type(file_name)[0] or (
            "video/mp4" if item["kind"] == VehicleMedia.Kind.VIDEO else "image/jpeg"
        )
        media = VehicleMedia.objects.create(
            vehicle=vehicle,
            kind=item["kind"],
            url=url,
            thumbnail_url=url if item["kind"] == VehicleMedia.Kind.PHOTO else None,
            content_type=content_type,
            file_name=file_name,
            s3_key=f"external/{vehicle.id}/{item['kind']}/{sort_order + 1}",
            original_url=url,
            status=VehicleMedia.Status.READY,
            sort_order=sort_order,
        )
        if first_photo is None and item["kind"] == VehicleMedia.Kind.PHOTO:
            first_photo = media
    if first_photo is not None:
        vehicle.cover_media = first_photo
        vehicle.save(update_fields=["cover_media", "updated_at"])


def _read_tabular_rows(uploaded_file) -> list[list[Any]]:
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
    if not raw:
        raise ValidationError("Uploaded file is empty.")

    if name.endswith(".csv") or name.endswith(".txt"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [
            list(row)
            for row in reader
            if any(str(cell).strip() for cell in row)
            and not str(row[0]).lstrip().startswith("#")
        ]

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            values = list(row)
            if any(cell is not None and str(cell).strip() for cell in values):
                rows.append(values)
        workbook.close()
        return rows

    raise ValidationError("Upload a .csv or .xlsx file.")


def build_sample_csv_bytes() -> bytes:
    buffer = io.StringIO()
    for instruction in CSV_OPTION_GUIDE:
        buffer.write(f"{instruction}\n")
    writer = csv.DictWriter(buffer, fieldnames=BULK_UPLOAD_COLUMNS)
    writer.writeheader()
    writer.writerows(SAMPLE_ROWS)
    return buffer.getvalue().encode("utf-8")


def build_sample_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(BULK_UPLOAD_COLUMNS)
    for row in SAMPLE_ROWS:
        sheet.append([row.get(column, "") for column in BULK_UPLOAD_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(BULK_UPLOAD_COLUMNS)).coordinate}"
    for field, options in FIELD_OPTIONS.items():
        column_index = BULK_UPLOAD_COLUMNS.index(field) + 1
        column_letter = sheet.cell(1, column_index).column_letter
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=False,
        )
        validation.error = f"Choose one of: {', '.join(options)}"
        validation.errorTitle = f"Invalid {field}"
        validation.prompt = f"Choose: {', '.join(options)}"
        validation.promptTitle = field
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}1001")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_vehicles_from_upload(*, request, dealer, uploaded_file) -> dict[str, Any]:
    rows = _read_tabular_rows(uploaded_file)
    if len(rows) < 2:
        raise ValidationError("File must include a header row and at least one vehicle row.")

    header_map = _map_headers(rows[0])
    location = (
        getattr(request.user, "preferred_location", None)
        if getattr(request.user, "preferred_location", None)
        and request.user.preferred_location.dealer_id == dealer.id
        else None
    )
    if location is None:
        location = DealerLocation.objects.filter(dealer=dealer, is_primary=True).first()
    if location is None:
        location = DealerLocation.objects.filter(dealer=dealer).first()
    if location is None:
        raise ValidationError("Add a dealer stand before bulk uploading vehicles.")

    defaults = {"locationId": str(location.id)}
    listing_limit = get_listing_limit(dealer)
    media_limits = get_media_limits(dealer)
    created: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    skipped_limit = 0

    for offset, row in enumerate(rows[1:], start=2):
        if listing_limit is not None and active_listing_count(dealer) + len(created) >= listing_limit:
            skipped_limit += 1
            continue
        try:
            payload, media_links = _row_to_payload(row, header_map, defaults=defaults)
            photo_count = sum(
                item["kind"] == VehicleMedia.Kind.PHOTO for item in media_links
            )
            video_count = sum(
                item["kind"] == VehicleMedia.Kind.VIDEO for item in media_links
            )
            if photo_count > media_limits["photosPerVehicle"]:
                raise ValidationError(
                    f"Your plan allows up to {media_limits['photosPerVehicle']} photos per car."
                )
            if video_count > media_limits["videosPerVehicle"]:
                raise ValidationError(
                    f"Your plan allows up to {media_limits['videosPerVehicle']} videos per car."
                )
            with transaction.atomic():
                serializer = VehicleSerializer(
                    data=payload,
                    context={"request": request},
                )
                serializer.is_valid(raise_exception=True)
                vehicle = serializer.save(dealer=dealer)
                _create_linked_media(vehicle, media_links)
            created.append(VehicleSerializer(vehicle, context={"request": request}).data)
        except (ValidationError, SerializerValidationError) as exc:
            detail = getattr(exc, "detail", str(exc))
            if isinstance(detail, dict):
                message = "; ".join(f"{key}: {value}" for key, value in detail.items())
            elif isinstance(detail, list):
                message = "; ".join(str(item) for item in detail)
            else:
                message = str(detail)
            failed.append({"row": offset, "error": message})
        except Exception as exc:  # noqa: BLE001 - surface row-level failures to the client
            failed.append({"row": offset, "error": str(exc)})

    return {
        "created": created,
        "count": len(created),
        "failed": failed,
        "failedCount": len(failed),
        "skippedForListingLimit": skipped_limit,
        "totalRows": len(rows) - 1,
    }
